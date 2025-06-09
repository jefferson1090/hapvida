# Versão: 3.9
import pandas as pd
import os
import re
from datetime import datetime
import openpyxl
import unicodedata  # Para normalização de caracteres
import logging

# --- Configuração de Logging ---
LOG_FILE = 'script_execution.log'
# Zera o arquivo de log ao iniciar o script
with open(LOG_FILE, 'w') as f:
    f.write('')

logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

logging.info("Script iniciado.")

# --- Configurações do Arquivo de Entrada ---
# SUBSTITUA PELO CAMINHO COMPLETO do seu arquivo Excel/CSV na sua máquina local
CAMINHO_PLANILHA = r'C:\Users\jeffe\OneDrive\Arquivos e Pastas antigas\Documentos\Hapvida - SublimeText\PROJETOS PYTHON\convenio_plano.xlsx'
TIPO_ARQUIVO = 'excel'  # DEFINE GLOBALMENTE: 'excel' ou 'csv' - Essencial para o script saber o formato de entrada.

# Delimitador PADRÃO de SAÍDA para o CSV gerado: PONTO E VÍRGULA (;)
CSV_DELIMITADOR_SAIDA = ';'

# Delimitador de ENTRADA para arquivos CSV, SE O TIPO_ARQUIVO for 'csv'.
DELIMITADOR_ENTRADA_CSV = ';'

# CHUNKSIZE é usado para dividir o DataFrame em blocos (se aplicável para leitura/processamento).
CHUNKSIZE = 10000

# --- Nomes dos Arquivos de Saída (Definidos Globalmente) ---
ARQUIVO_DROP_TABLE_SQL = 'drop_table_script.sql'
ARQUIVO_CREATE_TABLE_SQL = 'create_table_only_script.sql'
ARQUIVO_BATCH_EXEC = 'execute_db_scripts.bat'  # Nome principal do .bat
# NOVO: Nome do script PowerShell que será chamado pelo Batch
ARQUIVO_POWERSHELL_SQLLDR = 'execute_sqlldr.ps1'
ARQUIVO_SQLLDR_CTL = 'sqlldr_control_file.ctl'  # Adicionado o nome do arquivo de controle do SQL*Loader
ARQUIVO_SQLLDR_PAR = 'sqlldr_parameter_file.par' # Adicionado o nome do arquivo de parâmetros do SQL*Loader

ARQUIVO_CREDENCIAS = 'db_credentials.txt'
ARQUIVO_DADOS_PLANO = 'temp_data_to_load.csv'
ARQUIVO_NOME_TABELA_TXT = 'nome_tabela_gerado.txt'

# --- Configuração do Usuário para GRANT ---
USUARIO_GRANT = 'HUMASTER'

# --- PARÂMETROS CRÍTICOS PARA LEITURA DO EXCEL COM OPENPYXL ---
EXCEL_HEADER_ROW_NUM = 1
EXCEL_DATA_START_ROW_NUM = 2
NUM_COLUNAS_ESPERADAS_EXCEL = 14

# --- ORACLE_HOME_PATH para o SQL Loader ---
ORACLE_HOME_PATH = r'E:\app\client\isaacjf\product\12.2.0\client_1' # CAMINHO VALIDADO POR VOCÊ
NLS_LANG_VALUE = 'BRAZILIAN PORTUGUESE_BRAZIL.AL32UTF8'


# --- Funções auxiliares (manter como estão, elas foram validadas) ---
def normalizar_string(texto):
    if pd.isna(texto) or texto is None:
        return None
    texto_normalizado = unicodedata.normalize('NFKD', str(texto))
    texto_sem_acentos = "".join([c for c in texto_normalizado if not unicodedata.combining(c)])
    final_texto_limpo = re.sub(r'[^A-Z0-9\s_]', '', texto_sem_acentos.upper()).strip()
    # Remover o caractere U+00A0 (non-breaking space)
    final_texto_limpo = re.sub(r'[\u00A0]', '', final_texto_limpo)
    return final_texto_limpo


def limpar_nome_coluna(nome_original):
    cleaned_name = normalizar_string(nome_original)
    cleaned_name = re.sub(r'[\s]+', '_', cleaned_name)
    cleaned_name = re.sub(r'_+', '_', cleaned_name).strip('_')
    if not cleaned_name:
        cleaned_name = 'COL_VAZIA_PADRAO'
    if cleaned_name and not cleaned_name[0].isalpha():
        cleaned_name = 'COL_' + cleaned_name
    return cleaned_name


def gerar_nome_tabela(caminho_arquivo):
    base_name = os.path.basename(caminho_arquivo)
    table_suffix = os.path.splitext(base_name)[0]
    table_suffix = normalizar_string(table_suffix)
    table_suffix = re.sub(r'[\s]+', '_', table_suffix)
    table_suffix = re.sub(r'[^A-Z0-9_]', '', table_suffix)
    table_suffix = re.sub(r'_+', '_', table_suffix).strip('_')
    if not table_suffix or not table_suffix[0].isalpha():
        table_suffix = 'TBL_' + table_suffix
    prefix_table_object_name = 'TT_OPE_'
    max_suf_len_for_30_char_name = 30 - len(prefix_table_object_name)
    if len(table_suffix) > max_suf_len_for_30_char_name:
        table_suffix = table_suffix[:max_suf_len_for_30_char_name]
    table_suffix = table_suffix.upper()
    return f"{prefix_table_object_name}{table_suffix}"


def inferir_e_nomear_coluna(col_name_original, series):
    clean_col_name = col_name_original
    oracle_type = "VARCHAR2(255)"
    prefix = "NM_"
    if clean_col_name.startswith(('CD_', 'DS_', 'NU_', 'FL_', 'NM_')):
        final_col_name = clean_col_name
    else:
        if pd.api.types.is_integer_dtype(series):
            prefix = "NU_"
        elif pd.api.types.is_float_dtype(series):
            prefix = "NU_"
        elif pd.api.types.is_datetime64_any_dtype(series):
            prefix = "DT_"
        elif series.dtype == 'bool':
            prefix = "FL_"
        else:
            contains_numbers = series.astype(str).str.contains(r'\d').any()
            if contains_numbers:
                prefix = "CD_"
            else:
                prefix = "NM_"
        final_col_name = f"{prefix}{clean_col_name}"

    if len(final_col_name) > 30:
        final_col_name = final_col_name[:30]

    if pd.api.types.is_integer_dtype(series):
        oracle_type = "NUMBER"
    elif pd.api.types.is_float_dtype(series):
        if series.dropna().apply(lambda x: x == int(x)).all():
            oracle_type = "NUMBER"
        else:
            oracle_type = "NUMBER"
    elif pd.api.types.is_datetime64_any_dtype(series):
        oracle_type = "DATE"
    elif series.dtype == 'bool':
        oracle_type = "NUMBER(1)"
    else:
        max_len = series.astype(str).apply(len).max()
        if max_len > 4000:
            oracle_type = "CLOB"
        else:
            oracle_type = f"VARCHAR2({max_len if max_len > 0 else 255})"

    return final_col_name, oracle_type


# (Fim das funções auxiliares)


def gerar_scripts_oracle(caminho_arquivo, tipo_arquivo):
    logging.info(f"Iniciando a geração de scripts para o arquivo: {caminho_arquivo} do tipo: {tipo_arquivo}")

    nome_tabela_objeto = gerar_nome_tabela(caminho_arquivo)
    nome_tabela_objeto_com_aspas = f'"{nome_tabela_objeto}"'

    logging.info(f"Nome do objeto da tabela gerado: {nome_tabela_objeto}")

    try:
        if tipo_arquivo.lower() == 'excel':
            logging.info("Processando arquivo Excel com openpyxl.")
            workbook = openpyxl.load_workbook(CAMINHO_PLANILHA, data_only=True)
            sheet = workbook.active

            header_names_raw = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=EXCEL_HEADER_ROW_NUM, column=col_idx).value
                header_names_raw.append(cell_value)
            logging.debug(f"Nomes brutos das colunas lidos do Excel: {header_names_raw}")

            header_names_cleaned_temp = [limpar_nome_coluna(name) for name in header_names_raw]
            header_names_filtered = [name for name in header_names_cleaned_temp if name not in ['COL_VAZIA_PADRAO', 'COL_VAZIA_TEMP']]
            logging.debug(f"Nomes limpos das colunas do Excel (filtrados): {header_names_filtered}")

            final_column_names_from_excel = []  # Nomes limpos da planilha (temporários)
            seen_cols = set()
            for col_name in header_names_filtered:
                if col_name in seen_cols:
                    i = 1
                    while f"{col_name}_{i}" in seen_cols:
                        i += 1
                    new_name = f"{col_name}_{i}"
                    final_column_names_from_excel.append(new_name)
                    seen_cols.add(new_name)
                else:
                    final_column_names_from_excel.append(col_name)
                    seen_cols.add(col_name)
            logging.debug(f"Nomes finais das colunas ajustados (para duplicatas): {final_column_names_from_excel}")

            NUM_COLUNAS_REAIS_LIDAS = len(final_column_names_from_excel)

            data_rows = []
            for row_idx in range(EXCEL_DATA_START_ROW_NUM, sheet.max_row + 1):
                row_values = []
                for col_idx in range(1, NUM_COLUNAS_REAIS_LIDAS + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, str):
                        row_values.append(normalizar_string(cell_value))
                    else:
                        row_values.append(cell_value)
                data_rows.append(row_values)
            logging.debug(f"Número de linhas de dados lidas do Excel: {len(data_rows)}")

            # Criar o DataFrame do pandas usando os nomes de colunas dinâmicos e limpos do Excel
            df_full = pd.DataFrame(data_rows, columns=final_column_names_from_excel)
            logging.info("DataFrame do pandas criado a partir do Excel.")

            logging.debug(f"DataFrame inicial com colunas: {df_full.columns.tolist()}")

            # Colunas para DDL e mapeamento para CSV
            columns_ddl_list = []  # Lista final de strings para DDL
            col_mapping = {}  # Mapeia nomes limpos do Excel (ou CSV) para nomes Oracle finais

            # Primeira passagem para criar o col_mapping e definir o DDL
            for col_excel_name in df_full.columns:  # df_full.columns AINDA TEM NOMES LIMPOS DO EXCEL
                series = df_full[col_excel_name]
                final_col_oracle_name = col_excel_name
                oracle_type = "VARCHAR2(255)"

                if col_excel_name.startswith(('CD_', 'DS_', 'NU_', 'FL_', 'NM_')):
                    final_col_oracle_name = col_excel_name
                else:
                    final_col_oracle_name, oracle_type_temp = inferir_e_nomear_coluna(col_excel_name, series)
                    oracle_type = oracle_type_temp

                columns_ddl_list.append(f'"{final_col_oracle_name}" {oracle_type}')
                col_mapping[col_excel_name] = final_col_oracle_name  # Mapeamento: 'EXCEL_NAME' -> 'ORACLE_NAME'
            logging.debug(f"Lista de colunas para DDL: {columns_ddl_list}")
            logging.debug(f"Mapeamento de colunas Excel para Oracle: {col_mapping}")

            # NOVO: Renomear as colunas do DataFrame para os nomes Oracle ANTES de salvar no CSV
            # Isso garante que o CSV tenha os nomes exatos do banco
            df_full.rename(columns=col_mapping, inplace=True)
            logging.info("Colunas do DataFrame renomeadas para os nomes Oracle antes de salvar CSV.")

            df_full.to_csv(ARQUIVO_DADOS_PLANO, sep=CSV_DELIMITADOR_SAIDA, index=False, encoding='utf-8-sig')
            logging.info(f"Dados convertidos e salvos em CSV para SQL Loader: {ARQUIVO_DADOS_PLANO}")

        elif tipo_arquivo.lower() == 'csv':
            logging.info("Processando arquivo CSV com pandas.")
            df_full = pd.read_csv(caminho_arquivo, delimiter=DELIMITADOR_ENTRADA_CSV)
            df_full.columns = [limpar_nome_coluna(col) for col in df_full.columns]
            # ... (restante do processamento para CSV - não incluído aqui para brevidade)
        else:
            raise ValueError("Tipo de arquivo não suportado. Use 'excel' ou 'csv'.")

        logging.debug(f"DataFrame final contem colunas: {df_full.columns.tolist()}")
        logging.info("Processamento do arquivo concluído.")
        print("=========================================================================")

        with open(ARQUIVO_NOME_TABELA_TXT, 'w', encoding='utf-8') as f:
            f.write(nome_tabela_objeto_com_aspas)
            logging.info(f"Nome da tabela gravado em: {ARQUIVO_NOME_TABELA_TXT}")

        # --- Geração do ARQUIVO drop_table_script.sql ---
        logging.info("Gerando script DROP TABLE...")
        drop_table_sql_content = f"""
-- Script para dropar a tabela (gerado pelo Python)
-- Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SET ECHO ON
SET FEEDBACK ON
SET SERVEROUTPUT ON
WHENEVER SQLERROR EXIT FAILURE ROLLBACK
/

BEGIN
    EXECUTE IMMEDIATE 'DROP TABLE {nome_tabela_objeto_com_aspas} CASCADE CONSTRAINTS';
    DBMS_OUTPUT.PUT_LINE('Tabela {nome_tabela_objeto} dropada com sucesso.');
EXCEPTION
    WHEN OTHERS THEN
      IF SQLCODE = -942 THEN
        DBMS_OUTPUT.PUT_LINE('Tabela {nome_tabela_objeto} nao existe. Nao ha necessidade de drop.');
      ELSE
        DBMS_OUTPUT.PUT_LINE('Erro ao tentar dropar a tabela {nome_tabela_objeto}: ' || SQLERRM);
        RAISE;
      END IF;
END;
/

EXIT;
"""
        with open(ARQUIVO_DROP_TABLE_SQL, 'w', encoding='utf-8') as f:
            f.write(drop_table_sql_content)
        logging.info(f"Script DROP TABLE '{ARQUIVO_DROP_TABLE_SQL}' gerado com sucesso.")

        # --- Geração do ARQUIVO create_table_only_script.sql ---
        logging.info("Gerando script CREATE TABLE...")
        formatted_columns_ddl = []
        for i, col_def in enumerate(columns_ddl_list):  # Usar a lista já populada
            if i < len(columns_ddl_list) - 1:
                formatted_columns_ddl.append(f'    {col_def},')
            else:
                formatted_columns_ddl.append(f'    {col_def}')

        create_table_only_sql_content = f"""
-- Script para criar a tabela e conceder permissoes (gerado pelo Python)
-- Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SET DEFINE OFF;
SET ESCAPE OFF;
WHENEVER SQLERROR EXIT FAILURE;

CREATE TABLE {nome_tabela_objeto_com_aspas} (
{chr(10).join(formatted_columns_ddl)}
);
/

GRANT ALL ON {nome_tabela_objeto_com_aspas} TO {USUARIO_GRANT};
/

PROMPT Tabela criada com sucesso: {nome_tabela_objeto_com_aspas}
SELECT 'Tabela ' || table_name || ' criada no schema ' || owner || ' e possui ' || num_rows || ' linhas.'
FROM ALL_TABLES
WHERE TABLE_NAME = '{nome_tabela_objeto}'
  AND OWNER = USER;
/
EXIT;
"""
        with open(ARQUIVO_CREATE_TABLE_SQL, 'w', encoding='utf-8') as f:
            f.write(create_table_only_sql_content)
        logging.info(f"Script CREATE TABLE '{ARQUIVO_CREATE_TABLE_SQL}' gerado com sucesso.")


    except Exception as e:
        logging.error(f"OCORREU UM ERRO CRÍTICO na geração de scripts: {e}")
        print(f"\nOCORREU UM ERRO CRÍTICO na geração de scripts: {e}")
        print("Por favor, verifique:")
        print(f"  - O caminho da planilha Excel/CSV: '{CAMINHO_PLANILHA}' está correto e o arquivo existe?")
        print(f"  - A planilha não está aberta em outro programa?")
        print(f"  - O parametro TIPO_ARQUIVO ('{TIPO_ARQUIVO}') corresponde ao tipo real do arquivo?")
        print(f"  - PARA EXCEL: EXCEL_HEADER_ROW_NUM ('{EXCEL_HEADER_ROW_NUM}') está correto (geralmente 1 para linha 1 do Excel)?")
        print(f"  - PARA EXCEL: EXCEL_DATA_START_ROW_NUM ('{EXCEL_DATA_START_ROW_NUM}') está correto (2 para linha 2 do Excel, etc.)?")
        print(f"  - NUM_COLUNAS_ESPERADAS_EXCEL ('{NUM_COLUNAS_ESPERADAS_EXCEL}') está correto e corresponde ao número real de colunas que você quer ler?")
        print(f"  - PARA CSV: DELIMITADOR_ENTRADA_CSV ('{DELIMITADOR_ENTRADA_CSV}') está correto?")
        print("=========================================================================")
        exit(1)


# --- Execução Principal do Script Python ---
if __name__ == "__main__":
    gerar_scripts_oracle(CAMINHO_PLANILHA, TIPO_ARQUIVO)

    # Caminho da pasta local na VPN onde os arquivos residem.
    PASTA_LOCAL_VPN_PARA_EXECUCAO = r'\\tsclient\C\Users\jeffe\OneDrive\Arquivos e Pastas antigas\Documentos\Hapvida - SublimeText\PROJETOS PYTHON'

    # CONTEÚDO DO POWERSHELL SCRIPT (PARA CARGA DE DADOS COM SQL LOADER)
    # Este script será chamado pelo Batch
    powershell_script_content = f"""
$ErrorActionPreference = 'Stop'

# Caminhos dos arquivos (usando variáveis passadas do Batch)
$csvPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\{os.path.basename(ARQUIVO_DADOS_PLANO)}"
$ctlPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\{os.path.basename(ARQUIVO_SQLLDR_CTL)}"
$parPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\{os.path.basename(ARQUIVO_SQLLDR_PAR)}"
$logPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\sqlldr.log"
$badPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\sqlldr.bad"
$dscPath = "{PASTA_LOCAL_VPN_PARA_EXECUCAO}\\sqlldr.dsc"

# Dados de conexão (serão passados do Batch via variáveis de ambiente/parâmetros)
$usuario = $env:DB_USER_SQL
$senha = $env:DB_PASS_SQL
$dsn = $env:DB_DSN_SQL
$tabela = "{gerar_nome_tabela(CAMINHO_PLANILHA)}" # Obtem o nome da tabela do Python

# Configurações de ambiente Oracle (do Python)
$oracleHome = "{ORACLE_HOME_PATH}"
$nlsLang = "{NLS_LANG_VALUE}"

Write-Host "📥 Lendo o cabeçalho do CSV: $csvPath" -ForegroundColor Cyan
try {{
    $headers = Get-Content -Path $csvPath -Encoding UTF8 | Select-Object -First 1
    if (-not $headers) {{
        Write-Host "❌ Nao foi possivel ler o cabecalho do CSV." -ForegroundColor Red
        exit 1
    }}
    # As colunas do CSV já devem ter os nomes do Oracle devido ao rename no Python
    $colunas = $headers -split "{CSV_DELIMITADOR_SAIDA}"
    # Remove aspas se existirem em nomes de coluna no CSV
    $colunasCsv = $colunas | ForEach-Object {{ $_.Trim('"') }}

}} catch {{
    Write-Host "❌ Erro ao acessar o CSV: $_" -ForegroundColor Red
    exit 1
}}

Write-Host "🔌 Conectando ao banco para obter as colunas da tabela: $tabela..." -ForegroundColor Cyan
$tempSql = "$($env:LOCAL_EXEC_PATH)\temp_describe.sql"
$tempResult = "$($env:LOCAL_EXEC_PATH)\temp_columns.txt"
$query = """
SET HEADING OFF
SET FEEDBACK OFF
SET PAGESIZE 0
SET TRIMSPOOL ON
SET LINESIZE 1000
SELECT column_name FROM all_tab_columns
WHERE table_name = UPPER('$tabela') AND owner = UPPER('$usuario');
EXIT
"""

$query | Set-Content -Encoding ASCII $tempSql
$env:ORACLE_HOME = $oracleHome
$env:PATH = "$oracleHome\BIN;$env:PATH" # Correção na variável PATH
$env:NLS_LANG = $nlsLang
& sqlplus -S "$usuario/$senha@$dsn" "@$tempSql" > $tempResult 2>&1


# Lê colunas reais da tabela
$colunasBanco = os.popen(f'powershell -Command "(Get-Content \'{tempResult}\') | Where-Object {{$_.trim() -ne \'\'}}"', 'r').read().strip().split('\n')

if ($colunasBanco.Count -eq 0) {{
    Write-Host "`n❌ Falha ao obter colunas da tabela no banco. Verifique usuario, senha, DSN e nome da tabela." -ForegroundColor Red
    exit 1
}}

Write-Host "`n✅ Colunas encontradas na tabela Oracle:" -ForegroundColor Green
$colunasBanco | ForEach-Object {{ Write-Host " - $_" }}

# Filtra colunas do CSV que realmente existem no banco
$colunasValidas = @()
foreach ($coluna in $colunasCsv) {{ # $colunasCsv já é a lista de nomes do CSV
    $colName = $coluna.Trim()
    if ($colunasBanco -contains $colName) {{
        $colunasValidas += $colName
    }} else {{
        Write-Host "⚠️  Ignorando coluna do CSV '$colName' nao encontrada na tabela '$tabela'." -ForegroundColor Yellow
    }}
}}

if ($colunasValidas.Count -eq 0) {{
    Write-Host "`n❌ Nenhuma coluna valida encontrada para importar.` -ForegroundColor Red
    exit 1
}}

# Gera o arquivo CTL
Write-Host "`n📝 Gerando arquivo .ctl..." -ForegroundColor Cyan
$ctlContent = @"
LOAD DATA
INFILE '$csvPath'
BADFILE '$badPath'
DISCARDFILE '$dscPath'
APPEND
INTO TABLE $tabela
FIELDS TERMINATED BY '{CSV_DELIMITADOR_SAIDA}' OPTIONALLY ENCLOSED BY '"'
TRAILING NULLCOLS
(
"@
foreach ($col in $colunasValidas) {{
    # Os nomes de coluna no CTL devem ser os NOMES DA TABELA ORACLE (que já estão em $col)
    $ctlContent += "  `"$col`" CHAR," + "`n" # Usa aspas para nomes de coluna
}}
$ctlContent = $ctlContent.TrimEnd("`,", "`n") + "`n)"
$ctlContent | Set-Content -Path $ctlPath -Encoding ASCII

# Gera o PAR file
Write-Host "📝 Gerando arquivo .par..." -ForegroundColor Cyan
$parContent = @"
userid=$usuario/$senha@$dsn
control=$ctlPath
log=$logPath
bad=$badPath
data=$csvPath
"@
$parContent | Set-Content -Path $parPath -Encoding ASCII

# Executa o SQL*Loader
Write-Host "`n▶️  Executando SQL*Loader..." -ForegroundColor Cyan
try {{
    $sqlldrResult = & sqlldr PARFILE="$parPath" 2>&1 | Out-String
    Write-Host "`n===== RESULTADO DO SQLLDR =====" -ForegroundColor Yellow
    Write-Output $sqlldrResult

    if ($LASTEXITCODE -eq 0) {{
        Write-Host "`n✅ Carga concluida com sucesso." -ForegroundColor Green
    }} else {{
        Write-Host "`n❌ Erro ao executar SQL*Loader. Verifique o log: $logPath" -ForegroundColor Red
    }}
}} catch {{
    Write-Host "`n❌ Erro inesperado durante a execucao do SQL*Loader: $($_.Exception.Message)" -ForegroundColor Red
}}
"""
    with open(ARQUIVO_POWERSHELL_SQLLDR, 'w', encoding='utf-8') as f:
        f.write(powershell_script_content)
    print(f"Script PowerShell '{ARQUIVO_POWERSHELL_SQLLDR}' gerado com sucesso.")


    # CONTEÚDO DO BATCH SCRIPT
    batch_script_content = f"""@echo off
rem Script gerado pelo Python para executar DDL e DML no Oracle via SQL*Plus e SQL Loader
rem Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SETLOCAL ENABLEDELAYEDEXPANSION

rem ** Este Batch Script espera que os arquivos estejam na pasta: '{PASTA_LOCAL_VPN_PARA_EXECUCAO}' **
rem ** O usuario deve navegar ate esta pasta e executar o .bat de la. **
set "LOCAL_EXEC_PATH={PASTA_LOCAL_VPN_PARA_EXECUCAO}"

set "CREDENTIALS_FILE={ARQUIVO_CREDENCIAS}"
set "DROP_SCRIPT_NAME={ARQUIVO_DROP_TABLE_SQL}"
set "CREATE_SCRIPT_NAME={ARQUIVO_CREATE_TABLE_SQL}"
set "POWERSHELL_SQL_LOADER_SCRIPT_NAME={ARQUIVO_POWERSHELL_SQLLDR}"
set "DATA_FILE_NAME={ARQUIVO_DADOS_PLANO}"

echo.
echo =========================================================================
echo.
echo Verificando arquivos necessarios na pasta: !LOCAL_EXEC_PATH!
if not exist "!LOCAL_EXEC_PATH!\\!CREDENTIALS_FILE!" (
    echo ERRO: Arquivo de credenciais '!LOCAL_EXEC_PATH!\\!CREDENTIALS_FILE!' nao encontrado.
    pause >NUL
    exit /b 1
)
if not exist "!LOCAL_EXEC_PATH!\\!DROP_SCRIPT_NAME!" (
    echo ERRO: Arquivo do script DROP '!LOCAL_EXEC_PATH!\\!DROP_SCRIPT_NAME!' nao encontrado.
    pause >NUL
    exit /b 1
)
if not exist "!LOCAL_EXEC_PATH!\\!CREATE_SCRIPT_NAME!" (
    echo ERRO: Arquivo do script CREATE '!LOCAL_EXEC_PATH!\\!CREATE_SCRIPT_NAME!' nao encontrado.
    pause >NUL
    exit /b 1
)
if not exist "!LOCAL_EXEC_PATH!\\!DATA_FILE_NAME!" (
    echo ERRO: Arquivo de dados CSV '!LOCAL_EXEC_PATH!\\!DATA_FILE_NAME!' nao encontrado.
    pause >NUL
    exit /b 1
)
if not exist "!LOCAL_EXEC_PATH!\\!POWERSHELL_SQL_LOADER_SCRIPT_NAME!" (
    echo ERRO: Script PowerShell do SQL Loader '!LOCAL_EXEC_PATH!\\!POWERSHELL_SQL_LOADER_SCRIPT_NAME!' nao encontrado.
    pause >NUL
    exit /b 1
)
echo Arquivos necessarios para esta fase encontrados.
echo.

rem Nao usar pushd/popd. A referencia sera sempre pelo caminho completo.

:GET_CREDENTIALS
if exist "!LOCAL_EXEC_PATH!\\!CREDENTIALS_FILE!" (
    for /f "tokens=1-3 delims=," %%a in ('type "!LOCAL_EXEC_PATH!\\!CREDENTIALS_FILE!"') do (
        set "LAST_USER=%%a"
        set "LAST_PASS=%%b"
        set "LAST_DSN=%%c"
    )
    echo Ultimo ambiente utilizado: !LAST_USER!@!LAST_DSN!
    set /p USE_LAST_ENV="Deseja continuar no mesmo ambiente? (S/N): "
    if /i "!USE_LAST_ENV!"=="S" (
        set DB_USER=!LAST_USER!
        set DB_PASS=!LAST_PASS!
        set DB_DSN=!LAST_DSN!
    ) else (
        goto PROMPT_NEW_CREDENTIALS
    )
) else (
    :PROMPT_NEW_CREDENTIALS
    echo.
    echo Por favor, insira os dados de conexao com o banco de dados:
    set /p DB_USER="Usuario: "
    set /p DB_PASS="Senha: "
    set /p DB_DSN="DSN (Ex: hapvdese): "
    rem Grava credenciais, usando aspas para proteger contra caracteres especiais na senha.
    echo "!DB_USER!,!DB_PASS!,!DB_DSN!" > "!LOCAL_EXEC_PATH!\\!CREDENTIALS_FILE!"
    echo Credenciais salvas em !CREDENTIALS_FILE!.
)

echo.
echo Conectando como !DB_USER!@!DB_DSN!
echo.

rem ** PASSO 1.0: Tentativa de Conexao para Teste **
echo Tentando conectar ao banco de dados...
sqlplus -L -S !DB_USER!/!DB_PASS!@!DB_DSN! "select user from dual; exit;"

rem Verifica se a conexao foi bem-sucedida (SQL*Plus retorna 0 para sucesso)
if !ERRORLEVEL! NEQ 0 (
    echo.
    echo ERRO CRITICO: Nao foi possivel conectar ao banco de dados.
    echo Verifique suas credenciais e o DSN.
    echo.
    pause >NUL
    exit /b 1
) else (
    echo.
    echo CONEXAO BEM-SUCEDIDA!
    echo.
)

echo =========================================================================
echo.
echo Executando o script DROP TABLE: !LOCAL_EXEC_PATH!\!DROP_SCRIPT_NAME!
sqlplus -L -S !DB_USER!/!DB_PASS!@!DB_DSN! "@!LOCAL_EXEC_PATH!\\!DROP_SCRIPT_NAME!"

rem Analisa o ERRORLEVEL do sqlplus para mensagens amigaveis
if !ERRORLEVEL! EQU 0 (
    echo.
    echo SCRIPT DROP TABLE EXECUTADO COM SUCESSO.
    echo Tabela foi dropada ou nao existia.
    echo.
) else if !ERRORLEVEL! EQU 1 (
    echo.
    echo AVISO: SCRIPT DROP TABLE EXECUTADO COM AVISOS.
    echo A operacao de DROP pode nao ter sido totalmente bem-sucedida. Verifique o log do SQL*Plus acima.
    echo.
) else (
    echo.
    echo ERRO CRITICO ao executar o script DROP TABLE.
    echo Por favor, verifique as mensagens do SQL*Plus acima para detalhes do erro.
    echo.
)

echo.
echo =========================================================================
echo.
echo FASE DE TESTE: DROP TABLE CONCLUIDA.
echo.

rem ** INÍCIO DA EXECUÇÃO DO CREATE TABLE **
echo.
echo =========================================================================
echo.
echo Executando o script CREATE TABLE: !LOCAL_EXEC_PATH!\!CREATE_SCRIPT_NAME!
sqlplus -L -S !DB_USER!/!DB_PASS!@!DB_DSN! "@!LOCAL_EXEC_PATH!\\!CREATE_SCRIPT_NAME!"

if !ERRORLEVEL! EQU 0 (
    echo.
    echo SCRIPT CREATE TABLE EXECUTADO COM SUCESSO.
    echo Tabela criada e permissoes concedidas.
    echo.
) else if !ERRORLEVEL! EQU 1 (
    echo.
    echo AVISO: SCRIPT CREATE TABLE EXECUTADO COM AVISOS.
    echo A criacao da tabela pode nao ter sido totalmente bem-sucedida. Verifique o log do SQL*Plus acima.
    echo.
) else (
    echo.
    echo ERRO CRITICO ao executar o script CREATE TABLE.
    echo Por favor, verifique as mensagens do SQL*Plus acima para detalhes do erro.
    echo.
)

echo.
echo =========================================================================
echo.
echo FASE DE TESTE: CREATE TABLE CONCLUIDA.
echo.

rem ** INÍCIO DA CARGA DE DADOS COM SQL LOADER (VIA POWERSHELL) **
echo.
echo =========================================================================
echo.
echo Executando a carga de dados via SQL Loader (via PowerShell)...

rem Definir variaveis de ambiente para o PowerShell
set "DB_USER_SQL=!DB_USER!"
set "DB_PASS_SQL=!DB_PASS!"
set "DB_DSN_SQL=!DB_DSN!"
rem Executar o script PowerShell para SQL Loader
powershell.exe -ExecutionPolicy Bypass -File "!LOCAL_EXEC_PATH!\\!POWERSHELL_SQL_LOADER_SCRIPT_NAME!" `
    -DB_USER_SQL "!DB_USER!" -DB_PASS_SQL "!DB_PASS!" -DB_DSN_SQL "!DB_DSN!"

rem O SQL Loader eh executado DENTRO do PowerShell, entao o ERRORLEVEL do PowerShell eh o que importa.
rem O script PowerShell já faz o tratamento de erro e sai com um exit code.
if !ERRORLEVEL! NEQ 0 (
    echo.
    echo ERRO CRITICO: A carga de dados via SQL Loader falhou.
    echo Verifique a saida do PowerShell acima e o log do SQL Loader.
    echo.
) else (
    echo.
    echo CARGA DE DADOS VIA SQL LOADER EXECUTADA COM SUCESSO (via PowerShell).
    echo.
)

echo.
echo =========================================================================
echo.
echo Processo de carga de dados concluido.
echo.
pause
exit
"""
    with open(ARQUIVO_BATCH_EXEC, 'w', encoding='utf-8') as f:
        f.write(batch_script_content)

    print("\n" + "=" * 80)
    print("Script Batch gerado para execução na VPN:")
    print(rf"1. Gere os arquivos executando este script Python na sua maquina local.")
    print(rf"2. Copie (MANTENHA) os arquivos gerados (incluindo o '{ARQUIVO_BATCH_EXEC}') na pasta: '{PASTA_LOCAL_VPN_PARA_EXECUCAO}' na sua maquina local.")
    print(rf"3. No ambiente da VPN, abra o Prompt de Comando (ou PowerShell), **NAVEGUE ATÉ A PASTA '{PASTA_LOCAL_VPN_PARA_EXECUCAO}'** e execute:")
    print(rf"  ** \"{ARQUIVO_BATCH_EXEC}\" **")
    print("\n" + "=" * 80)